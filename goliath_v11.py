#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════════╗
║                                                                                  ║
║    ██████╗  ██████╗ ██╗     ██╗ █████╗ ████████╗██╗  ██╗    ██╗   ██╗ █████╗    ║
║   ██╔════╝ ██╔═══██╗██║     ██║██╔══██╗╚══██╔══╝██║  ██║    ██║   ██║██╔══██╗   ║
║   ██║  ███╗██║   ██║██║     ██║███████║   ██║   ███████║    ██║   ██║╚█████╔╝   ║
║   ██║   ██║██║   ██║██║     ██║██╔══██║   ██║   ██╔══██║    ╚██╗ ██╔╝██╔══██╗   ║
║   ╚██████╔╝╚██████╔╝███████╗██║██║  ██║   ██║   ██║  ██║     ╚████╔╝ ╚█████╔╝   ║
║    ╚═════╝  ╚═════╝ ╚══════╝╚═╝╚═╝  ╚═╝   ╚═╝   ╚═╝  ╚═╝      ╚═══╝   ╚════╝    ║
║                                                                                  ║
║   CUPRA COMPETITION INTELLIGENCE ENGINE — v9.0 ULTRA                                ║
║   ═════════════════════════════════════════════                                   ║
║                                                                                  ║
║   ✦ Zero-Browser Engine (pure HTTP — 10× faster than Selenium)                   ║
║   ✦ Deep JSON Mining (__NEXT_DATA__ → full description, ALL parameters)          ║
║   ✦ Smart Memory — nie skanuje aut które już zna (cache.json)                    ║
║   ✦ Precision Margin Calculator z tabelą rabatów 2025/2026                       ║
║   ✦ Filtr: CUPRA only · ≥2025 · ≤30km · no SEAT/Ateca                           ║
║   ✦ Born & Tavascan → ELEKTRYK_OK (osobna logika marży TBD)                     ║
║   ✦ Incremental Excel (aktualizuje istniejący plik, nie nadpisuje)               ║
║   ✦ Multi-sheet XLSX: Inventory · Podsumowanie · Anomalie                        ║
║   ✦ Otomoto ID + Moc + Pojemność + Kolor + Skrzynia + Nadwozie                  ║
║   ✦ Wykrywa usunięte ogłoszenia (auta sprzedane/wycofane)                        ║
║   ✦ JSON export dla frontendu (snake_case, pełne dane)                           ║
║                                                                                  ║
║   Requirements: pip install requests openpyxl                                    ║
║                                                                                  ║
╚══════════════════════════════════════════════════════════════════════════════════╝
"""

import json
import logging
import os
import random
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional, List, Dict, Set, Tuple

import requests
from bs4 import BeautifulSoup

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


# ═══════════════════════════════════════════════════════════════════════════════
#  ⚙️  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

class Config:
    """Wszystkie ustawienia w jednym miejscu."""

    # --- Wydajność ---
    MAX_WORKERS: int = 5
    REQUEST_TIMEOUT: int = 25
    MAX_RETRIES: int = 3
    RETRY_DELAY: float = 3.0
    MIN_DELAY: float = 0.8
    MAX_DELAY: float = 2.0
    MAX_PAGES_PER_DEALER: int = 999  # Skanuj WSZYSTKIE strony!

    # --- Pliki wyjściowe ---
    OUTPUT_JSON: str = "data.json"
    OUTPUT_XLSX: str = "CUPRA_INVENTORY.xlsx"
    CACHE_FILE: str = "goliath_cache.json"
    LOG_FILE: str = "goliath.log"

    # --- Filtrowanie ---
    MIN_YEAR: int = 2024
    MAX_MILEAGE: int = 30         # Max przebieg w km (nowe auta)
    BLOCKED_MODELS: set = {"ateca", "cupra-ateca"}
    BLOCKED_MAKES: set = {"seat"}
    EV_MODELS: set = {"born", "tavascan", "cupra-born", "cupra-tavascan"}

    # --- Marża: progi kolorów ---
    MARGIN_GREAT: float = 4.0     # Zielony
    MARGIN_WARN: float = 2.0      # Żółty 0-2%
    MARGIN_BAD: float = 0.0       # Czerwony < 0%

    # --- Smart Memory ---
    # Marża OK = nie skanuj ponownie. Poza zakresem = zawsze skanuj.
    CACHE_OK_MARGIN_MIN: float = -0.6
    CACHE_OK_MARGIN_MAX: float = 7.0

    # --- Normalny zakres marży [-0.6%, 7.0%] ---
    MARGIN_NORMAL_MIN: float = -0.6
    MARGIN_NORMAL_MAX: float = 7.0


# ═══════════════════════════════════════════════════════════════════════════════
#  🏢  REJESTR DEALERÓW (29 salonów)
# ═══════════════════════════════════════════════════════════════════════════════

HOME_DEALER = "MOTORPOL WROCŁAW"  # 🏠 Nasz salon — punkt odniesienia

DEALERS = [
    {"name": "MOTORPOL WROCŁAW",     "url": "https://cupramotorpolwroclaw.otomoto.pl/inventory", "home": True},
    {"name": "PLICHTA GDYNIA",       "url": "https://cupra-plichta-gdynia.otomoto.pl/inventory"},
    {"name": "PLICHTA GDAŃSK",       "url": "https://cupra-plichta-gdansk.otomoto.pl/inventory"},
    {"name": "PLICHTA",              "url": "https://cupra-plichta.otomoto.pl/inventory"},
    {"name": "SZCZECIN",             "url": "https://cupraseat-szczecin.otomoto.pl/inventory"},
    {"name": "OLSZTYN",              "url": "https://cupra-olsztyn.otomoto.pl/inventory"},
    {"name": "CUPRA.OTOMOTO",        "url": "https://cupra.otomoto.pl/inventory"},
    {"name": "TORUŃ",                "url": "https://cupra-torun.otomoto.pl/inventory"},
    {"name": "BIACOMEX",             "url": "https://biacomexcupra.otomoto.pl/inventory"},
    {"name": "POL-CAR",              "url": "https://cupra-pol-car.otomoto.pl/inventory"},
    {"name": "GCZ POZNAŃ",           "url": "https://gcz-cupra-poznan.otomoto.pl/inventory"},
    {"name": "CARSED",               "url": "https://cupracarsed.otomoto.pl/inventory"},
    {"name": "WARSZAWA CENTRUM",     "url": "https://cupra-warszawa-centrum.otomoto.pl/inventory"},
    {"name": "KROTOSKI",             "url": "https://cupra-krotoski.otomoto.pl/inventory"},
    {"name": "GCZ WARSZAWA",         "url": "https://gcz-cupra-warszawa.otomoto.pl/inventory"},
    {"name": "WARSZAWA",             "url": "https://cupra-warszawa.otomoto.pl/inventory"},
    {"name": "CUPRA POLSKA",         "url": "https://cuprapolska.otomoto.pl/inventory"},
    {"name": "STUDIO ŁÓDŹ",          "url": "https://cuprastudiolodz.otomoto.pl/inventory"},
    {"name": "STUDIO",               "url": "https://cupra-studio.otomoto.pl/inventory"},
    {"name": "LUBIN",                "url": "https://cupralubin.otomoto.pl/inventory"},
    {"name": "CZĘSTOCHOWA",          "url": "https://cupra-czestochowa.otomoto.pl/inventory"},
    {"name": "KIELCE",               "url": "https://cuprakielce.otomoto.pl/inventory"},
    {"name": "LELLEK OPOLE",         "url": "https://seatlellekopole.otomoto.pl/inventory"},
    {"name": "LELLEK GLIWICE",       "url": "https://cupralellekgliwice.otomoto.pl/inventory"},
    {"name": "PRO-MOTO",             "url": "https://cupra-pro-moto.otomoto.pl/inventory"},
    {"name": "KRAKÓW",               "url": "https://cupra-krakow.otomoto.pl/inventory"},
    {"name": "STUDIO KRAKÓW",        "url": "https://cuprastudio-krakow.otomoto.pl/inventory"},
    {"name": "DYNAMICA",             "url": "https://cupra-dynamica.otomoto.pl/inventory"},
    {"name": "GG AUTO RZESZÓW",      "url": "https://cupraggautorzeszow.otomoto.pl/inventory"},
    {"name": "AUTO GAZDA",           "url": "https://cupraautogazda.otomoto.pl/inventory"},
]


# ═══════════════════════════════════════════════════════════════════════════════
#  💰  TABELA RABATÓW IMPORTERA
# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
# 💰 RABATY Z IMPORTERA — Wczytaj z rebates.json
# ═══════════════════════════════════════════════════════════════════════════════

def load_rebates():
    """Wczytaj rabaty i konfigurację z settings.json."""
    settings_file = "settings.json"
    if not os.path.exists(settings_file):
        logging.warning(f"  Brak {settings_file} — używam domyślnych rabatów")
        return [], {}, {}
    
    try:
        with open(settings_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        rebates = [r for r in data.get("rebates", []) if r.get("active", True)]
        margin_cfg = data.get("margin_config", {})
        dealer_pcts = data.get("dealer_percentages", {})
        
        logging.info(f"  Wczytane {len(rebates)} aktywnych rabatów z {settings_file}")
        return rebates, margin_cfg, dealer_pcts
    except Exception as e:
        logging.warning(f"  Błąd wczytywania {settings_file}: {e}")
        return [], {}, {}

REBATES_LIST, MARGIN_CFG, DEALER_PCTS = load_rebates()

# Override Config from settings.json
if MARGIN_CFG:
    Config.MARGIN_GREAT = MARGIN_CFG.get("margin_great", Config.MARGIN_GREAT)
    Config.MARGIN_WARN = MARGIN_CFG.get("margin_warn", Config.MARGIN_WARN)
    Config.MARGIN_BAD = MARGIN_CFG.get("margin_bad", Config.MARGIN_BAD)
    Config.MARGIN_NORMAL_MIN = MARGIN_CFG.get("margin_ok_min", -0.6)
    Config.MARGIN_NORMAL_MAX = MARGIN_CFG.get("margin_ok_max", 7.0)
    Config.CACHE_OK_MARGIN_MIN = MARGIN_CFG.get("cache_ok_margin_min", -0.6)
    Config.CACHE_OK_MARGIN_MAX = MARGIN_CFG.get("cache_ok_margin_max", 7.0)


# ═══════════════════════════════════════════════════════════════════════════════
#  📦  MODEL DANYCH
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class CarData:
    """Pełny rekord jednego samochodu."""

    # Identyfikacja
    otomoto_id: str = ""
    dealer: str = ""
    dealer_short: str = ""
    title: str = ""
    model: str = ""                 # Formentor / Leon / Terramar / Born / Tavascan
    model_raw: str = ""             # Surowa wartość z Otomoto (cupra-formentor)
    year: int = 0
    fuel: str = ""
    power_hp: int = 0
    engine_capacity: str = ""       # np. "1 498 cm3"
    mileage_km: int = 0
    gearbox: str = ""
    drive: str = ""
    color: str = ""
    color_type: str = ""            # Metalik / Mat / Perła
    body_type: str = ""
    doors: int = 0
    seats: int = 0
    is_new: bool = True
    has_vin: bool = False

    # Ceny
    sale_price: int = 0
    catalog_price: int = 0
    catalog_price_from_desc: bool = False  # True = znaleziona w opisie
    price_30d: int = 0

    # Kalkulacja marży
    dealer_cost: int = 0
    rebate: int = 0
    rebate_applied: bool = False    # True = rabat importera został faktycznie użyty
    force_no_rebate: bool = False   # v10: User override — wyłącz rabat (cena kat == sprzedaży)
    margin_pln: int = 0
    margin_pct: float = 0.0

    # Typ pojazdu
    vehicle_type: str = "new"       # "new" / "demo" / "used"
    is_demo: bool = False           # True = demo/ekspozycyjne/testowe/short-term

    margin_without_rebate_pct: float = 0.0  # Marża BEZ rabatu
    margin_with_rebate_pct: float = 0.0     # Marża Z rabatem (aktualna)
    rebate_from_description: bool = False    # Czy rabat znaleziony w opisie
    price_after_rebate_from_desc: int = 0   # Cena po rabacie z opisu
    special_edition: str = ""                # Tribe Edition, Limited, etc.
    note: str = ""                           # Notatka (short demo, itd.)

    # Opis dealera
    description: str = ""

    # Meta
    has_catalog_price: bool = False
    is_ev: bool = False             # Born / Tavascan
    status: str = ""
    url: str = ""
    scraped_at: str = ""
    first_seen: str = ""            # Data pierwszego skanowania
    price_changed: bool = False     # Czy cena się zmieniła od ostatniego skanu
    is_home: bool = False           # True = nasz salon (MOTORPOL WROCŁAW)

    def to_dict(self) -> dict:
        """Konwersja do dict dla JSON."""
        return {
            "otomoto_id": self.otomoto_id,
            "dealer": self.dealer,
            "dealer_short": self.dealer_short,
            "title": self.title,
            "model": self.model,
            "model_raw": self.model_raw,
            "year": self.year,
            "fuel": self.fuel,
            "power_hp": self.power_hp,
            "engine_capacity": self.engine_capacity,
            "mileage_km": self.mileage_km,
            "gearbox": self.gearbox,
            "drive": self.drive,
            "color": self.color,
            "color_type": self.color_type,
            "body_type": self.body_type,
            "doors": self.doors,
            "seats": self.seats,
            "is_new": self.is_new,
            "has_vin": self.has_vin,
            "sale_price": self.sale_price,
            "catalog_price": self.catalog_price,
            "price_30d": self.price_30d,
            "dealer_cost": self.dealer_cost,
            "rebate": self.rebate,
            "rebate_applied": self.rebate_applied,
            "force_no_rebate": self.force_no_rebate,
            "margin_pln": self.margin_pln,
            "margin_pct": self.margin_pct,
            "vehicle_type": self.vehicle_type,
            "is_demo": self.is_demo,
            "description": self.description,
            "has_catalog_price": self.has_catalog_price,
            "is_ev": self.is_ev,
            "status": self.status,
            "url": self.url,
            "otomoto_id": self.otomoto_id,
            "scraped_at": self.scraped_at,
            "first_seen": self.first_seen,
            "rebate_from_description": self.rebate_from_description,
            "price_after_rebate_from_desc": self.price_after_rebate_from_desc,
            "margin_without_rebate_pct": self.margin_without_rebate_pct,
            "margin_with_rebate_pct": self.margin_with_rebate_pct,
            "special_edition": self.special_edition,
            "note": self.note,
            "price_changed": self.price_changed,
            "is_home": self.is_home,
        }


# ═══════════════════════════════════════════════════════════════════════════════
#  🧠  SMART MEMORY (Cache System)
# ═══════════════════════════════════════════════════════════════════════════════

class SmartMemory:
    """
    Inteligentna pamięć skanera. Zapamiętuje wyniki i:
    - Pomija auta ze stabilną marżą (nie marnuje requestów)
    - Zawsze skanuje ponownie auta z anomaliami lub brakiem ceny
    - Wykrywa usunięte ogłoszenia
    - Śledzi historię cen (zmiana ceny → re-scan)
    """

    def __init__(self, cache_file: str = Config.CACHE_FILE):
        self.cache_file = cache_file
        self.cache: Dict[str, dict] = {}   # key = URL
        self._load()

    def _load(self):
        """Wczytaj cache z pliku."""
        if os.path.exists(self.cache_file):
            try:
                with open(self.cache_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    # Konwertuj listę do dict po URL
                    for item in data:
                        url = item.get("url", "")
                        if url:
                            self.cache[url] = item
                elif isinstance(data, dict):
                    self.cache = data
                logging.info(f"  🧠 Pamięć wczytana: {len(self.cache)} aut w cache")
            except Exception as e:
                logging.warning(f"  ⚠️ Błąd cache: {e} — startuję od zera")
                self.cache = {}

    def save(self, cars: List[CarData]):
        """Zapisz wyniki do cache."""
        cache_data = {}
        for car in cars:
            cache_data[car.url] = {
                "url": car.url,
                "otomoto_id": car.otomoto_id,
                "sale_price": car.sale_price,
                "catalog_price": car.catalog_price,
                "margin_pct": car.margin_pct,
                "status": car.status,
                "vehicle_type": car.vehicle_type,
                "scraped_at": car.scraped_at,
                "first_seen": car.first_seen,
            }
        self.cache = cache_data
        with open(self.cache_file, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
        logging.info(f"  🧠 Cache zapisany: {len(cache_data)} rekordów")

    def should_skip(self, url: str) -> bool:
        """
        Czy pominąć to ogłoszenie?
        Pomijamy TYLKO gdy:
          - Jest w cache
          - Ma status OK lub ELEKTRYK_OK
          - Marża w bezpiecznym zakresie
        """
        if url not in self.cache:
            return False

        entry = self.cache[url]
        status = entry.get("status", "")
        margin = entry.get("margin_pct", None)

        # Zawsze skanuj ponownie jeśli status nie jest OK (strict: only OK)
        if status != "OK":
            return False

        # v10.0: MEGA FIX — zawsze RESCAN auta z anomaliami marży
        if margin is not None:
            # Marża >6% = zawsze RESCAN (rabat może być błędny)
            # Marża <-0.5% = zawsze RESCAN (demo/używane do sprawdzenia)
            if margin > 6.0 or margin < -0.5:
                return False  # Force RESCAN
            # Marża w zakresie [-0.5%, 6.0%] = OK, pomiń
            if not (Config.CACHE_OK_MARGIN_MIN <= margin <= Config.CACHE_OK_MARGIN_MAX):
                return False

        return True

    def get_first_seen(self, url: str) -> str:
        """Zwróć datę pierwszego skanowania lub teraz."""
        if url in self.cache:
            return self.cache[url].get("first_seen", "")
        return ""

    def detect_price_change(self, url: str, new_sale_price: int) -> bool:
        """Sprawdź czy cena się zmieniła."""
        if url in self.cache:
            old_price = self.cache[url].get("sale_price", 0)
            if old_price > 0 and old_price != new_sale_price:
                return True
        return False

    def get_removed_urls(self, active_urls: Set[str]) -> List[dict]:
        """Znajdź ogłoszenia które zniknęły z Otomoto (sprzedane/wycofane)."""
        removed = []
        for url, entry in self.cache.items():
            if url not in active_urls:
                removed.append(entry)
        return removed


# ═══════════════════════════════════════════════════════════════════════════════
#  🌐  KLIENT HTTP
# ═══════════════════════════════════════════════════════════════════════════════

class HttpClient:
    """Thread-safe HTTP client z retry i anti-bot headerami."""

    USER_AGENTS = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Safari/605.1.15",
    ]

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "pl-PL,pl;q=0.9,en-US;q=0.8,en;q=0.7",
            "Accept-Encoding": "gzip, deflate, br",
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
        })
        self._request_count = 0

    def get(self, url: str) -> Optional[str]:
        for attempt in range(1, Config.MAX_RETRIES + 1):
            try:
                self.session.headers["User-Agent"] = random.choice(self.USER_AGENTS)
                delay = random.uniform(Config.MIN_DELAY, Config.MAX_DELAY)
                time.sleep(delay)

                resp = self.session.get(url, timeout=Config.REQUEST_TIMEOUT, allow_redirects=True)
                self._request_count += 1

                if resp.status_code == 200:
                    return resp.text
                elif resp.status_code == 429:
                    wait = Config.RETRY_DELAY * (2 ** attempt)
                    logging.warning(f"  ⚠️  HTTP 429 — czekam {wait:.0f}s...")
                    time.sleep(wait)
                elif resp.status_code in (403, 503):
                    wait = Config.RETRY_DELAY * attempt
                    logging.warning(f"  ⚠️  HTTP {resp.status_code} — czekam {wait:.0f}s (próba {attempt}/{Config.MAX_RETRIES})")
                    time.sleep(wait)
                else:
                    logging.warning(f"  ⚠️  HTTP {resp.status_code}: {url}")
                    break

            except requests.exceptions.Timeout:
                logging.warning(f"  ⏳ Timeout (próba {attempt}/{Config.MAX_RETRIES})")
                time.sleep(Config.RETRY_DELAY * attempt)
            except requests.exceptions.ConnectionError as e:
                logging.warning(f"  🔌 Connection error (próba {attempt}/{Config.MAX_RETRIES})")
                time.sleep(Config.RETRY_DELAY * attempt)
            except requests.RequestException as e:
                logging.error(f"  ❌ Request error: {e}")
                break

        return None

    @property
    def total_requests(self) -> int:
        return self._request_count


# ═══════════════════════════════════════════════════════════════════════════════
#  🔗  KOLEKTOR LINKÓW Z INVENTORY (z paginacją)
# ═══════════════════════════════════════════════════════════════════════════════

class InventoryCollector:
    """Zbiera linki do ofert z dealer inventory pages."""

    # Zbiera WSZYSTKIE /oferta/ linki - relative lub absolute
    OFFER_LINK_RE = re.compile(
        r'href="([^"]*(?:/oferta/)[^"?#]*)"'
    )

    def __init__(self, client: HttpClient):
        self.client = client

    def _extract_ads_from_json(self, html: str) -> list:
        """Wyciąga ogłoszenia z __NEXT_DATA__ JSON (urqlState → publishedAds)."""
        try:
            match = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
            if not match:
                return [], 0
            
            data = json.loads(match.group(1))
            urql_state = data.get('props', {}).get('pageProps', {}).get('urqlState', {})
            
            for key in urql_state:
                entry = urql_state[key]
                if isinstance(entry, dict) and 'data' in entry:
                    parsed = json.loads(entry['data']) if isinstance(entry['data'], str) else entry['data']
                    if 'publishedAds' in parsed:
                        ads = parsed['publishedAds'].get('ads', [])
                        total = parsed['publishedAds'].get('total', 0)
                        return ads, total
            return [], 0
        except Exception as e:
            logging.debug(f"    ⚠️ JSON extraction error: {e}")
            return [], 0

    def collect(self, dealer_url: str) -> Set[str]:
        all_links: Set[str] = set()
        base_url = dealer_url.split("?")[0]

        for page in range(1, Config.MAX_PAGES_PER_DEALER + 1):
            url = f"{base_url}?page={page}" if page > 1 else base_url
            html = self.client.get(url)
            if not html:
                logging.debug(f"    🔴 Strona {page}: BRAK HTML (timeout/block)")
                break

            # 🔧 v13: Multi-strategy link extraction
            ads, total = self._extract_ads_from_json(html)
            
            page_links = set()
            
            if ads:
                for ad in ads:
                    ad_url = ad.get('url', '')
                    if ad_url and '/oferta/' in ad_url:
                        page_links.add(ad_url)
                logging.debug(f"    📄 Strona {page}: JSON → {len(page_links)} linków (total w systemie: {total})")
            
            # Fallback 1: BeautifulSoup (always run if JSON gave few results)
            if len(page_links) < 3:
                try:
                    soup = BeautifulSoup(html, 'html.parser')
                    for a in soup.find_all('a', href=True):
                        href = a['href'].strip()
                        if '/oferta/' in href:
                            page_links.add(href)
                    if page_links:
                        logging.debug(f"    📄 Strona {page}: BS4 fallback → {len(page_links)} linków")
                except Exception as e:
                    logging.debug(f"    ⚠️ BS4 fallback error: {e}")
            
            # Fallback 2: Regex on raw HTML (last resort)
            if len(page_links) < 3:
                regex_links = self.OFFER_LINK_RE.findall(html)
                for href in regex_links:
                    if '/oferta/' in href:
                        page_links.add(href)
                if regex_links:
                    logging.debug(f"    📄 Strona {page}: Regex fallback → {len(page_links)} linków")

            # Pre-filter na poziomie URL
            filtered = set()
            for link in page_links:
                link_lower = link.lower()
                if "/oferta/seat-" in link_lower and "cupra" not in link_lower:
                    continue
                if "ateca" in link_lower:
                    continue
                link = link.rstrip("\\")
                filtered.add(link)

            new_links = filtered - all_links
            logging.debug(f"    🔍 Po filtrach = {len(filtered)}, Nowych = {len(new_links)}")
            
            if not new_links:
                logging.info(f"    ✅ Strona {page}: Brak nowych - koniec skanowania")
                break

            all_links.update(new_links)

            if page > 1:
                logging.info(f"    📄 Strona {page}: +{len(new_links)} nowych")

        if all_links:
            logging.debug(f"    ✔️ Razem z dealera: {len(all_links)} linków")
        return all_links


# ═══════════════════════════════════════════════════════════════════════════════
#  🧮  KALKULATOR MARŻY
# ═══════════════════════════════════════════════════════════════════════════════

class MarginCalculator:
    """
    v9.0 ULTRA — Smart Rebate Logic:

    1. Oblicz marżę BEZ rabatu:
       dealer_cost_base = Cena Katalogowa × 0.94 (−6%)

    2. Jeśli marża bez rabatu jest w normalnym zakresie [-0.5%, 6.0%]:
       → NIE stosuj rabatu (marża i tak jest zdrowa)

    3. Jeśli marża jest POZA zakresem:
       → Zastosuj rabat importera i przelicz

    Efekt: Rabat koryguje anomalie, a nie tworzy sztuczne zawyżenie marży.
    """

    @staticmethod
    def get_fuel_type(description: str, title: str) -> str:
        """Rozpoznaj typ paliwa z opisu i tytułu."""
        text = f"{description} {title}".lower()
        
        # PHEV / Plug-in
        if any(x in text for x in ["phev", "plug-in", "plugin", "spalinowo-elektryczny"]):
            return "PHEV"
        # Diesel
        if "diesel" in text or "d5" in text:
            return "diesel"
        # Benzyna + Elektryk (obydwa)
        if "elektryk" in text and ("benzyna" in text or "lpg" in text):
            return "benzyna_elektryk"
        # Elektryk
        if "elektryk" in text or "ev" in text:
            return "elektryk"
        # LPG
        if "lpg" in text:
            return "lpg"
        # Domyślnie benzyna
        return "benzyna"
    
    # Mapowanie wariantów modelu na klucz rabatu (Leon ST = Leon, itd.)
    MODEL_REBATE_MAP = {
        "leon st": "leon",
        "leon-sportstourer": "leon",
        "leon sportstourer": "leon",
    }

    @staticmethod
    def get_rebate(year: int, model_raw: str, title: str, fuel: str = "benzyna") -> int:
        """
        Dopasuj rabat z settings.json.
        Priorytet: standardowe > promocyjne.
        Dla VZ modeli: szukaj specjalnych rabatów is_vz.
        """
        model_key = model_raw.lower().replace("cupra-", "").replace("cupra ", "")
        # Leon ST / Leon Sportstourer → "leon" (te same rabaty)
        model_key = MarginCalculator.MODEL_REBATE_MAP.get(model_key, model_key)
        title_lower = title.lower()
        
        # Detect VZ variant
        is_vz = any(x in title_lower for x in ['vz ', ' vz', 'vz,'])
        
        best_rebate = 0
        best_priority = -1
        
        for entry in REBATES_LIST:
            if not entry.get("active", True):
                continue

            # Skip percentage-based entries (handled separately in calculate_v10)
            if entry.get("extra_pct", 0) > 0:
                continue
                
            entry_model = entry.get("model", "").lower()
            if entry_model != model_key:
                continue
            
            entry_year = entry.get("year", 0)
            if entry_year != year:
                continue
            
            # Fuel matching
            entry_fuel = entry.get("fuel", "all").lower()
            fuel_lower = fuel.lower()
            
            if entry_fuel != "all":
                # EV matching
                if entry_fuel == "elektryk" and fuel_lower not in ("elektryk", "elektryczny", "ev"):
                    continue
                elif entry_fuel == "phev" and fuel_lower not in ("phev", "plug-in", "plugin", "spalinowo-elektryczny"):
                    continue
                elif entry_fuel == "benzyna" and fuel_lower in ("elektryk", "ev", "phev", "plug-in"):
                    continue
            
            # VZ matching
            entry_is_vz = entry.get("is_vz", False)
            if entry_is_vz and not is_vz:
                continue
            if is_vz and entry_is_vz:
                # VZ match gets highest priority
                priority = 10
            elif not entry_is_vz:
                priority = 5
            else:
                continue
            
            # Prefer standardowe over promocyjne
            if entry.get("type", "") == "standardowe":
                priority += 2
            elif entry.get("type", "") == "promocyjne":
                priority += 1
            
            amount = entry.get("amount", 0)
            
            if priority > best_priority or (priority == best_priority and amount > best_rebate):
                best_rebate = amount
                best_priority = priority
        
        return best_rebate

    @staticmethod
    def calculate(catalog_price: int, sale_price: int, year: int,
                  model_raw: str, title: str) -> Tuple[int, int, float, int, bool]:
        """Returns: (dealer_cost, rebate_used, margin_pct, margin_pln, rebate_applied)"""
        rebate_available = MarginCalculator.get_rebate(year, model_raw, title)
        TARGET_MIN = -1.0
        TARGET_MAX = 7.0

        # KROK 1: Oblicz BEZ rabatu
        dealer_cost_base = int(catalog_price * 0.94)
        margin_pln_base = sale_price - dealer_cost_base
        margin_pct_base = round((margin_pln_base / sale_price) * 100, 2) if sale_price > 0 else 0.0

        # KROK 2: Czy marża jest w zakresie?
        if TARGET_MIN <= margin_pct_base <= TARGET_MAX:
            return dealer_cost_base, 0, margin_pct_base, margin_pln_base, False

        # KROK 3: Marża PONIŻEJ zakresu → rabat może pomóc
        if margin_pct_base < TARGET_MIN and rebate_available > 0:
            dealer_cost = dealer_cost_base - rebate_available
            margin_pln = sale_price - dealer_cost
            margin_pct = round((margin_pln / sale_price) * 100, 2) if sale_price > 0 else 0.0
            if TARGET_MIN <= margin_pct <= TARGET_MAX:
                return dealer_cost, rebate_available, margin_pct, margin_pln, True

        # Fallback: najlepsze co mamy (bez fikcyjnych korekt VGP)
        if rebate_available > 0:
            dealer_cost = dealer_cost_base - rebate_available
            margin_pln = sale_price - dealer_cost
            margin_pct = round((margin_pln / sale_price) * 100, 2) if sale_price > 0 else 0.0
            return dealer_cost, rebate_available, margin_pct, margin_pln, True

        return dealer_cost_base, 0, margin_pct_base, margin_pln_base, False

    @staticmethod
    def calculate_v10(catalog_price: int, sale_price: int, year: int,
                      model_raw: str, title: str, description: str = "",
                      rebate_from_desc: int = 0,
                      price_after_rebate_from_desc: int = 0) -> Tuple[int, int, float, int, bool, float, float]:
        """
        v12 — CORRECTED MARGIN CALCULATION:
        
        Base discount = korekta_vgp from settings.json:
          Leon/Born/Tavascan: 13%
          Formentor/Terramar: 9%
        
        Logic:
          1. Calculate margin WITHOUT PLN rebate (but WITH korekta_vgp discount)
          2. If in range [-0.6%, 7%] and NOT EV → DONE (dealer already priced in rebate)
          3. If outside range OR if EV model → apply PLN rebate and recalculate
          4. Born/Tavascan (EV): ALWAYS apply PLN rebate
        
        Returns: (dealer_cost, rebate_used, margin_pct, margin_pln, rebate_applied,
                  margin_without_rebate_pct, margin_with_rebate_pct)
        """
        model_key = model_raw.lower().replace("cupra-", "").replace("cupra ", "")
        # Leon ST / Leon Sportstourer → "leon" (same rebates and dealer percentages)
        model_key = MarginCalculator.MODEL_REBATE_MAP.get(model_key, model_key)
        
        ev_models = {"born", "tavascan"}
        is_ev = model_key in ev_models
        
        # Force elektryk fuel for Born/Tavascan
        if is_ev:
            fuel = "elektryk"
        else:
            fuel = MarginCalculator.get_fuel_type(description, title)
        
        # ── Dealer discount from VGP (korekta) ──
        # This is the % the dealer gets off catalog price from the importer
        # Leon/Born/Tavascan: 13%, Formentor/Terramar: 9%
        model_pcts = DEALER_PCTS.get(model_key, {})
        korekta_pct = model_pcts.get("korekta_vgp", 13.0) / 100.0
        
        # ── Base dealer cost (WITH korekta but WITHOUT PLN rebate) ──
        dealer_cost_base = int(catalog_price * (1 - korekta_pct))
        
        # ── Find best PLN rebate from importer ──
        rebate_available = MarginCalculator.get_rebate(year, model_raw, title, fuel)
        
        # Check for rebate mentioned in dealer description
        if rebate_from_desc > 0:
            rebate_available = max(rebate_available, rebate_from_desc)
        
        # ── Handle percentage-based rebates (e.g., Tavascan 19%, VZ 24%+18500) ──
        title_lower = title.lower()
        is_vz = any(x in title_lower for x in ['vz ', ' vz', 'vz,', ' vz.'])
        
        extra_pct_rebate = 0
        extra_pln_from_pct_entry = 0
        for entry in REBATES_LIST:
            entry_model = entry.get("model", "").lower()
            entry_is_vz = entry.get("is_vz", False)
            if (entry_model == model_key
                    and entry.get("year") == year
                    and entry.get("extra_pct", 0) > 0
                    and entry.get("active", True)):
                # VZ matching — skip mismatches
                if entry_is_vz and not is_vz:
                    continue  # Skip VZ entry for non-VZ car
                if not entry_is_vz and is_vz:
                    continue  # Skip non-VZ entry for VZ car
                extra_pct_rebate = int(catalog_price * entry["extra_pct"] / 100)
                extra_pln_from_pct_entry = entry.get("amount", 0)
                break
        
        # ── Calculate total rebate (pick best option) ──
        if extra_pct_rebate > 0:
            # Percentage rebate available (e.g., 19% or 24%+18500)
            pct_total = extra_pct_rebate + extra_pln_from_pct_entry
            total_rebate = max(rebate_available, pct_total)
        else:
            total_rebate = rebate_available
        
        # ── Step 1: Margin WITHOUT PLN rebate ──
        margin_pln_base = sale_price - dealer_cost_base
        margin_pct_base = round((margin_pln_base / sale_price) * 100, 2) if sale_price > 0 else 0.0
        margin_without_rebate = margin_pct_base
        
        # ── Step 2: Calculate margin WITH rebate (for comparison/display) ──
        if total_rebate > 0:
            dealer_cost_with = dealer_cost_base - total_rebate
            margin_pln_with = sale_price - dealer_cost_with
            margin_pct_with = round((margin_pln_with / sale_price) * 100, 2) if sale_price > 0 else 0.0
        else:
            dealer_cost_with = dealer_cost_base
            margin_pln_with = margin_pln_base
            margin_pct_with = margin_pct_base
        
        margin_with_rebate = margin_pct_with
        
        # ── Step 3: AGGRESSIVE MARGIN FIT — force into [-1%, 7%] by any means ──
        TARGET_MIN = -1.0
        TARGET_MAX = 7.0
        
        def _try_margin(cost, reb, applied):
            """Calculate margin for given cost/rebate combo."""
            m_pln = sale_price - cost
            m_pct = round((m_pln / sale_price) * 100, 2) if sale_price > 0 else 0.0
            return cost, reb, m_pct, m_pln, applied, margin_without_rebate, margin_with_rebate
        
        # Strategy 1: Base margin (no rebate) — check if already in range
        if TARGET_MIN <= margin_pct_base <= TARGET_MAX:
            return _try_margin(dealer_cost_base, 0, False)
        
        # Strategy 2: Apply full rebate
        if total_rebate > 0 and TARGET_MIN <= margin_pct_with <= TARGET_MAX:
            return _try_margin(dealer_cost_with, total_rebate, True)
        
        # Strategy 3: EV models — always apply rebate (even if out of range)
        if is_ev and total_rebate > 0:
            return _try_margin(dealer_cost_with, total_rebate, True)
        
        # Strategy 4: FALLBACK — zawsze używaj PRAWDZIWEGO kosztu dealera!
        # ⚠️ USUNIĘTO Strategie 4 i 5 które wymyślały fikcyjne korekty VGP (6%-20%).
        # Przez to koszt dealera był za niski/wysoki a marże przekłamane.
        # Teraz: dealer_cost = zawsze catalog × (1 - korekta_vgp z settings.json)
        if total_rebate > 0:
            # Wybierz wersję (z/bez rabatu) bliżej zakresu docelowego
            dist_base = min(abs(margin_pct_base - TARGET_MIN), abs(margin_pct_base - TARGET_MAX))
            dist_with = min(abs(margin_pct_with - TARGET_MIN), abs(margin_pct_with - TARGET_MAX))
            if dist_with <= dist_base:
                return _try_margin(dealer_cost_with, total_rebate, True)
        
        return _try_margin(dealer_cost_base, 0, False)


# ═══════════════════════════════════════════════════════════════════════════════
#  🔬  PARSER OGŁOSZENIA (Deep JSON Mining)
# ═══════════════════════════════════════════════════════════════════════════════

class OfferParser:
    """
    Parsuje stronę ogłoszenia Otomoto.
    v9.0: Dodano wykrywanie demo/używanych z tytułu i opisu.
    """

    NEXT_DATA_RE = re.compile(
        r'<script\s+id="__NEXT_DATA__"[^>]*>(.*?)</script>', re.S
    )

    # ── Regex na cenę katalogową ──
    # Obsługuje WSZYSTKIE formaty dealerów:
    #   "Cena katalogowa: 153 900 zł"
    #   "Cena katalogowa pojazdu: 236 716,00 PLN brutto"
    #   "Cena katalogowa: 240 661,00 zł"
    #   "cena cennikowa 153.900 PLN"
    #   "katlogowa: 153 900 zł" (typo)
    CATALOG_PRICE_RE = re.compile(
        r'(?:cena\s+katalogowa|cena\s+cennikowa|cena\s+z\s+cennika|'
        r'cena\s+konfiguracji|'                    # CARSED format
        r'warto[śs][ćc]\s+katalogowa|cena\s+przed\s+rabatem|'
        r'cena\s+regularna|cena\s+standardowa|'
        r'katalogowa|cennikowa|katlogowa)'       # Keyword
        r'[^0-9]{0,30}?'                         # Max 30 non-digit chars
        r'([\d][\d\s.,]{3,})',                    # Price digits
        re.I
    )

    # ── Regex na "Cena po rabacie" i "Rabat" (v10.0) ──
    PRICE_AFTER_REBATE_RE = re.compile(
        r'cena\s+(?:pojazdu\s+)?po\s+rabat(?:cie|em)?'
        r'[^0-9]{0,30}?'
        r'([\d][\d\s.,]{3,})',
        re.I
    )
    
    REBATE_AMOUNT_RE = re.compile(
        r'rabat(?:em|\s+całkowit)?(?:y)?[^0-9]{0,30}?'
        r'([\d][\d\s.,]{3,})',
        re.I
    )

    # ── Regex na cenę z 30 dni ──
    PRICE_30D_RE = re.compile(
        r'(?:najni[żz]sza\s+cena|cena\s+z\s+ostatnich\s+30|'
        r'ostatnich\s+30\s+dni|30\s+dni\s+przed)'
        r'[^0-9]{0,40}?'
        r'([\d][\d\s.,]{3,})',
        re.I
    )

    # ── Słowa kluczowe: DEMO / UŻYWANE / SHORT-TERM ──
    DEMO_KEYWORDS = [
        'demonstracyjn', 'demo ', ' demo', 'ex-demo', 'ex demo',
        'testow', 'jazd testow', 'auto z jazd', 'pojazd z jazd',
        'ekspozycyjn', 'wystawow', 'z ekspozycji',
        'short term', 'short-term', 'krótkoterminow',
        'przedpremierow',
        'wyprzeda',                       # "Wyprzedaż DEMO" (KRAKÓW)
        'zarejestrowany na dealer',       # registered to dealer = demo (AUTO GAZDA)
        'zarejestrowane na dealer',
        'samochód z salonu',              # exhibition car
    ]

    USED_KEYWORDS = [
        'używan', 'poleasingow', 'po leasingu',
        'second hand', 'z drugiej ręki',
    ]

    def __init__(self, client: HttpClient, memory: SmartMemory):
        self.client = client
        self.memory = memory

    def _detect_vehicle_type(self, title: str, description: str,
                              is_new_param: bool) -> Tuple[str, bool]:
        """
        Wykryj typ pojazdu na podstawie tytułu, opisu i parametru new/used.
        Returns: (vehicle_type, is_demo)
        """
        search_text = f"{title} {description}".lower()

        # Sprawdź słowa demo/ekspozycyjne
        for keyword in self.DEMO_KEYWORDS:
            if keyword in search_text:
                return "demo", True

        # Sprawdź słowa używane
        for keyword in self.USED_KEYWORDS:
            if keyword in search_text:
                return "used", False

        # Parametr Otomoto: is_new=False → używane
        if not is_new_param:
            return "used", False

        return "new", False

    def parse(self, url: str, dealer_short: str) -> Optional[CarData]:
        """Pobierz i sparsuj stronę ogłoszenia."""
        html = self.client.get(url)
        if not html:
            return None

        match = self.NEXT_DATA_RE.search(html)
        if not match:
            logging.debug(f"  Brak __NEXT_DATA__: {url}")
            return None

        try:
            next_data = json.loads(match.group(1))
            ad = next_data["props"]["pageProps"]["advert"]
        except (json.JSONDecodeError, KeyError, TypeError):
            return None

        # ── Details dict ──
        details: Dict[str, str] = {}
        for d in ad.get("details", []):
            details[d.get("key", "")] = d.get("value", "")

        # ── parametersDict (bardziej kompletne) ──
        params_dict = ad.get("parametersDict", {})

        def get_param_value(key: str) -> str:
            """Wyciągnij wartość z parametersDict."""
            entry = params_dict.get(key, {})
            vals = entry.get("values", [])
            if vals:
                return str(vals[0].get("value", ""))
            return ""

        def get_param_label(key: str) -> str:
            """Wyciągnij label z parametersDict."""
            entry = params_dict.get(key, {})
            vals = entry.get("values", [])
            if vals:
                return str(vals[0].get("label", ""))
            return ""

        # ── Filtrowanie: marka ──
        make = details.get("make", "").strip().lower()
        if make in Config.BLOCKED_MAKES:
            return None

        # ── Filtrowanie: model ──
        model_raw = get_param_value("model") or details.get("model", "")
        model_raw_lower = model_raw.strip().lower()
        if model_raw_lower in Config.BLOCKED_MODELS:
            return None

        # Clean model name
        model_display = details.get("model", model_raw).strip()
        if model_display.lower().startswith("cupra "):
            model_display = model_display[6:]
        # Normalize model names
        model_name_map = {
            "leon": "Leon",
            "leon-sportstourer": "Leon ST",
            "leon sportstourer": "Leon ST",
            "formentor": "Formentor",
            "terramar": "Terramar",
            "born": "Born",
            "tavascan": "Tavascan",
        }
        model_display = model_name_map.get(model_display.lower().strip(), model_display)

        # ── Filtrowanie: rok ──
        year_str = get_param_value("year") or details.get("year", "0")
        year = int(year_str) if year_str.isdigit() else 0
        if year < Config.MIN_YEAR:
            return None

        # ── Filtrowanie: przebieg (max 30km = nowe) ──
        mileage_str = get_param_value("mileage")
        mileage = int(re.sub(r'[^\d]', '', mileage_str) or 0)
        if mileage > Config.MAX_MILEAGE:
            return None

        # ── Opis (HTML → czysty tekst) ──
        description_html = ad.get("description", "")
        clean_desc = re.sub(r'<[^>]+>', ' ', description_html)
        clean_desc = re.sub(r'\s+', ' ', clean_desc)

        # ── Cena sprzedaży ──
        price_data = ad.get("price", {})
        sale_price = int(price_data.get("value", 0))


        # v11: Nie filtrujemy roku 2024 (Tavascan MY24 potrzebny)
        
        # ── Cena katalogowa (z opisu) ──
        catalog_price = 0
        cat_match = self.CATALOG_PRICE_RE.search(clean_desc)
        if cat_match:
            catalog_price = self._parse_price(cat_match.group(1))
            # Sanity check: cena kat. powinna być w rozsądnym zakresie
            # Min 50k PLN, Max 800k PLN (Born od ~130k, Tavascan ~300k+)
            # Jeśli poza zakresem → prawdopodobnie typo w ogłoszeniu
            if catalog_price < 50_000 or catalog_price > 800_000:
                logging.debug(f"  ⚠ Cena katalogowa {catalog_price} poza zakresem, ignoruję")
                catalog_price = 0

        # ── Cena 30 dni ──
        price_30d = 0
        p30_match = self.PRICE_30D_RE.search(clean_desc)
        if p30_match:
            price_30d = self._parse_price(p30_match.group(1))


        # ── Cena po rabacie i Rabat (z opisu) — v10.0 ──
        price_after_rebate_from_desc = 0
        rebate_from_desc = 0
        
        par_match = self.PRICE_AFTER_REBATE_RE.search(clean_desc)
        if par_match:
            price_after_rebate_from_desc = self._parse_price(par_match.group(1))
        
        rab_match = self.REBATE_AMOUNT_RE.search(clean_desc)
        if rab_match:
            rebate_from_desc = self._parse_price(rab_match.group(1))
        
        # ── Parametry pojazdu ──
        power_label = get_param_label("engine_power")
        power = int(re.sub(r'[^\d]', '', power_label) or 0)

        engine_cap_label = get_param_label("engine_capacity")
        fuel_label = get_param_label("fuel_type")
        gearbox_label = get_param_label("gearbox")
        drive_label = get_param_label("transmission")
        color_label = get_param_label("color")
        color_type_label = get_param_label("colour_type")
        body_type_label = get_param_label("body_type")

        doors_str = get_param_value("door_count")
        doors = int(doors_str) if doors_str.isdigit() else 0

        seats_str = get_param_value("nr_seats")
        seats = int(seats_str) if seats_str.isdigit() else 0

        new_used = get_param_value("new_used")
        has_vin = get_param_value("has_vin") == "1"

        # ── EV detection ──
        is_ev = model_raw_lower in Config.EV_MODELS or any(
            ev in model_raw_lower for ev in ("born", "tavascan")
        )

        # ── Demo/Used detection (v9.0) ──
        title_text = ad.get("title", "")
        vehicle_type, is_demo = self._detect_vehicle_type(
            title_text, clean_desc, (new_used == "new")
        )


        # ── Special Edition Detection (v10.0) ──
        special_edition = ""
        special_keywords = ["tribe edition", "special edition", "limited edition", "limited", "tribe", "s edition"]
        search_for_edition = title_text.lower() + " " + clean_desc.lower()
        for keyword in special_keywords:
            if keyword in search_for_edition:
                special_edition = keyword.replace(" edition", "").replace(" ", "").title()
                if "tribe" in keyword.lower():
                    special_edition = "Tribe Edition"
                elif "limited" in keyword.lower():
                    special_edition = "Limited"
                break
        
        # ── Dodatkowa heurystyka demo: is_new=False w parametrach ──
        if vehicle_type == "new" and new_used != "new":
            vehicle_type = "used"  # Otomoto mówi 'used' → szanujemy

        # ── Seller ──
        seller = ad.get("seller", {})
        dealer_full = seller.get("name", dealer_short)

        # ── Otomoto ID ──
        otomoto_id = str(ad.get("id", ""))
        if not title_text:
            title_text = ad.get("title", "")

        # ── First seen tracking ──
        first_seen = self.memory.get_first_seen(url) or datetime.now().strftime("%Y-%m-%d %H:%M")

        # ── Price change detection ──
        price_changed = self.memory.detect_price_change(url, sale_price)

        # ── Detect REAL fuel type (PHEV / Benzyna / Diesel / Elektryk) ──
        fuel_real = MarginCalculator.get_fuel_type(clean_desc, title_text)

        # ── Build CarData ──
        car = CarData(
            otomoto_id=otomoto_id,
            dealer=dealer_full,
            dealer_short=dealer_short,
            title=ad.get("title", ""),
            model=model_display,
            model_raw=model_raw,
            year=year,
            fuel=fuel_real,
            power_hp=power,
            engine_capacity=engine_cap_label,
            mileage_km=mileage,
            gearbox=gearbox_label,
            drive=drive_label,
            color=color_label,
            color_type=color_type_label,
            body_type=body_type_label,
            doors=doors,
            seats=seats,
            is_new=(new_used == "new"),
            has_vin=has_vin,
            sale_price=sale_price,
            catalog_price=catalog_price,
            catalog_price_from_desc=(catalog_price > 0),
            price_30d=price_30d,
            has_catalog_price=(catalog_price > 0),
            description=clean_desc[:2000],  # Increased from 500 to preserve catalog price info
            is_ev=is_ev,
            vehicle_type=vehicle_type,
            is_demo=is_demo,
            price_after_rebate_from_desc=price_after_rebate_from_desc,
            rebate_from_description=(rebate_from_desc > 0),
            special_edition=special_edition,
            note="",  # Zostanie ustawiony przy kalkulacji marży
            url=url,
            scraped_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
            first_seen=first_seen,
            price_changed=price_changed,
            is_home=(dealer_short == HOME_DEALER),
        )

        # ── Margin calculation ──
        if is_ev:
            # Born & Tavascan — pełna logika marży z rabatami EV (v11)
            if car.has_catalog_price and car.sale_price > 0:
                car.dealer_cost, car.rebate, car.margin_pct, car.margin_pln, car.rebate_applied, margin_without_rebate, margin_with_rebate = (
                    MarginCalculator.calculate_v10(
                        car.catalog_price, car.sale_price, car.year,
                        car.model_raw, car.title, car.description,
                        rebate_from_desc, price_after_rebate_from_desc
                    )
                )
                car.margin_without_rebate_pct = margin_without_rebate
                car.margin_with_rebate_pct = margin_with_rebate
                
                # Status for EVs — nigdy ANOMALIA (elektryki mają inne struktury VGP)
                if vehicle_type == "demo":
                    car.status = "DEMO"
                elif vehicle_type == "used":
                    car.status = "UŻYWANE"
                elif Config.MARGIN_NORMAL_MIN <= car.margin_pct <= Config.MARGIN_NORMAL_MAX:
                    car.status = "OK"
                else:
                    car.status = "ELEKTRYK_OK"  # Born/Tavascan zawsze ELEKTRYK_OK, nie ANOMALIA"
            else:
                car.status = "ELEKTRYK_OK"
        elif car.has_catalog_price and car.sale_price > 0:
            # Oblicz marżę ze Smart Rebate (v10.0 PRO z obsługą paliwa)
            car.dealer_cost, car.rebate, car.margin_pct, car.margin_pln, car.rebate_applied, margin_without_rebate, margin_with_rebate = (
                MarginCalculator.calculate_v10(
                    car.catalog_price, car.sale_price, car.year,
                    car.model_raw, car.title, car.description,  # ← description dla rozpoznania paliwa!
                    rebate_from_desc, price_after_rebate_from_desc
                )
            )
            car.margin_without_rebate_pct = margin_without_rebate
            car.margin_with_rebate_pct = margin_with_rebate

            # --- Status classification (v10.0) ---
            if vehicle_type == "demo":
                car.status = "DEMO"
                car.note = ""
            elif vehicle_type == "used":
                car.status = "UŻYWANE"
                car.note = ""
            elif Config.MARGIN_NORMAL_MIN <= car.margin_pct <= Config.MARGIN_NORMAL_MAX:
                car.status = "OK"
                car.note = ""
            else:
                # Marża poza zakresem = ANOMALIA
                car.status = "ANOMALIA"
                
                # Notatka dla ujemnych marż (może to demo)
                if car.margin_pct < Config.MARGIN_NORMAL_MIN and not is_demo:
                    car.note = "⚠️ short demo / spad z uruchomioną gwarancją"
                elif special_edition:
                    car.note = f"✨ {special_edition}"
        else:
            car.status = "BRAK_CENY_KAT"

        return car


    @staticmethod
    def _parse_price(text: str) -> int:
        """
        Konwertuj polskie formaty cen do int:
          '153 900'      → 153900
          '153 900,00'   → 153900
          '236 716,00'   → 236716
          '180882'       → 180882
          '153.900'      → 153900
        """
        text = text.strip()

        # Usuń polską część dziesiętną: ",00" lub ".00" na końcu
        text = re.sub(r'[,\.]\s*\d{2}\s*$', '', text)

        # Usuń wszystko oprócz cyfr
        cleaned = re.sub(r'[^\d]', '', text)
        if not cleaned:
            return 0

        value = int(cleaned)

        # Sanity check — cena samochodu powinna być 50k-1.5M PLN
        if value < 50_000 or value > 1_500_000:
            return 0

        return value


# ═══════════════════════════════════════════════════════════════════════════════
#  💾  EKSPORTER (JSON + Multi-Sheet XLSX)
# ═══════════════════════════════════════════════════════════════════════════════

class Exporter:
    """Eksportuje wyniki do JSON (frontend) i XLSX (multi-sheet analityczny)."""

    # ── Kolory ──
    DARK_BLUE = "003366"
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if HAS_XLSX else None
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") if HAS_XLSX else None
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") if HAS_XLSX else None
    GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") if HAS_XLSX else None
    EV_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid") if HAS_XLSX else None

    @staticmethod
    def to_json(cars: List[CarData], filepath: str):
        cars_for_export = [
            car for car in cars 
            if car.status not in ("UŻYWANE", "BRAK_CENY_KAT") and car.vehicle_type != "used"
        ]
        # Export as clean array (dashboard expects list, not object)
        cars_data = [car.to_dict() for car in cars_for_export]
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(cars_data, f, ensure_ascii=False, indent=2)
        logging.info(f"  JSON -> {filepath} ({len(cars_for_export)} rekordow)")

    @staticmethod
    def to_xlsx(cars: List[CarData], filepath: str):
        if not HAS_XLSX:
            logging.warning("  ⚠️  Brak openpyxl — pomijam XLSX")
            return

        wb = openpyxl.Workbook()

        # ═══════════════════════════════════════
        #  Sheet 1: PEŁNY INVENTARZ
        # ═══════════════════════════════════════
        ws1 = wb.active
        ws1.title = "📋 Inventarz"

        headers = [
            "Dealer",           # A
            "Model",            # B
            "Pełny Tytuł",      # C
            "Rok",              # D
            "Paliwo",           # E
            "Moc KM",           # F
            "Pojemność",        # G
            "Przebieg km",      # H
            "Skrzynia",         # I
            "Napęd",            # J
            "Kolor",            # K
            "Typ koloru",       # L
            "Nadwozie",         # M
            "Cena Sprzedaży",   # N
            "Cena Katalogowa",  # O
            "Cena 30 dni",      # P
            "Rabat Importera",  # Q
            "Koszt Dealera",    # R
            "Marża PLN",        # S
            "Marża %",          # T
            "Status",           # U
            "Otomoto ID",       # V
            "Zmiana ceny?",     # W
            "Pierwszy skan",    # X
            "Ostatni skan",     # Y
            "URL",              # Z
        ]
        ws1.append(headers)

        for car in cars:
            ws1.append([
                car.dealer_short,
                car.model,
                car.title,
                car.year,
                car.fuel,
                car.power_hp,
                car.engine_capacity,
                car.mileage_km,
                car.gearbox,
                car.drive,
                car.color,
                car.color_type,
                car.body_type,
                car.sale_price,
                car.catalog_price if car.catalog_price > 0 else "",
                car.price_30d if car.price_30d > 0 else "",
                car.rebate if car.rebate > 0 else "",
                car.dealer_cost if car.dealer_cost > 0 else "",
                car.margin_pln if car.has_catalog_price else "",
                car.margin_pct if car.has_catalog_price else "",
                car.status,
                car.otomoto_id,
                "TAK" if car.price_changed else "",
                car.first_seen,
                car.scraped_at,
                car.url,
            ])

        Exporter._style_main_sheet(ws1, len(cars))

        # ═══════════════════════════════════════
        #  Sheet 2: PODSUMOWANIE DEALERÓW
        # ═══════════════════════════════════════
        ws2 = wb.create_sheet("📊 Podsumowanie")

        # Agreguj dane po dealerze
        dealer_stats: Dict[str, dict] = {}
        for car in cars:
            d = car.dealer_short
            if d not in dealer_stats:
                dealer_stats[d] = {
                    "total": 0, "with_margin_ok": 0, "without_margin": 0,
                    "ev": 0, "demo": 0,
                    "margins_ok": [],       # Tylko OK (do średniej)
                    "margins_all": [],      # Wszystkie z marżą
                    "sale_prices": [], "anomalies": 0,
                    "models": set()
                }
            s = dealer_stats[d]
            s["total"] += 1
            s["models"].add(car.model)
            s["sale_prices"].append(car.sale_price)
            if car.is_ev:
                s["ev"] += 1
            elif car.vehicle_type in ("demo", "used"):
                s["demo"] += 1
                if car.has_catalog_price:
                    s["margins_all"].append(car.margin_pct)
            elif car.has_catalog_price:
                s["margins_all"].append(car.margin_pct)
                if car.status == "OK":
                    s["with_margin_ok"] += 1
                    s["margins_ok"].append(car.margin_pct)
                elif car.status == "ANOMALIA":
                    s["anomalies"] += 1
            else:
                s["without_margin"] += 1

        ws2.append([
            "Dealer", "Aut ogółem", "Marża OK", "Demo/Używane",
            "Bez ceny kat.", "EV", "Anomalie",
            "Śr. marża % (OK)", "Min marża %", "Max marża %",
            "Śr. cena sprzedaży", "Modele"
        ])

        for dealer_name in sorted(dealer_stats.keys()):
            s = dealer_stats[dealer_name]
            margins_ok = s["margins_ok"]
            margins_all = s["margins_all"]
            prices = s["sale_prices"]
            ws2.append([
                dealer_name,
                s["total"],
                len(margins_ok),
                s["demo"],
                s["without_margin"],
                s["ev"],
                s["anomalies"],
                round(sum(margins_ok) / len(margins_ok), 2) if margins_ok else "",
                round(min(margins_all), 2) if margins_all else "",
                round(max(margins_all), 2) if margins_all else "",
                round(sum(prices) / len(prices)) if prices else "",
                ", ".join(sorted(s["models"])),
            ])

        Exporter._style_summary_sheet(ws2, len(dealer_stats))

        # ═══════════════════════════════════════
        #  Sheet 3: PODSUMOWANIE MODELI
        # ═══════════════════════════════════════
        ws3 = wb.create_sheet("🚗 Modele")

        model_stats: Dict[str, dict] = {}
        for car in cars:
            m = car.model or "Nieznany"
            if m not in model_stats:
                model_stats[m] = {
                    "total": 0, "margins_ok": [], "margins_all": [],
                    "prices": [], "years": set(), "dealers": set(),
                    "demo_count": 0
                }
            ms = model_stats[m]
            ms["total"] += 1
            ms["prices"].append(car.sale_price)
            ms["years"].add(car.year)
            ms["dealers"].add(car.dealer_short)
            if car.vehicle_type in ("demo", "used"):
                ms["demo_count"] += 1
            if car.has_catalog_price and not car.is_ev:
                ms["margins_all"].append(car.margin_pct)
                if car.status == "OK":
                    ms["margins_ok"].append(car.margin_pct)

        ws3.append([
            "Model", "Aut ogółem", "Demo/Używane", "Roczniki", "Dealerów",
            "Śr. cena", "Min cena", "Max cena",
            "Śr. marża % (OK)", "Min marża %", "Max marża %"
        ])

        for model_name in sorted(model_stats.keys()):
            ms = model_stats[model_name]
            margins_ok = ms["margins_ok"]
            margins_all = ms["margins_all"]
            prices = ms["prices"]
            ws3.append([
                model_name,
                ms["total"],
                ms["demo_count"],
                ", ".join(str(y) for y in sorted(ms["years"])),
                len(ms["dealers"]),
                round(sum(prices) / len(prices)) if prices else "",
                min(prices) if prices else "",
                max(prices) if prices else "",
                round(sum(margins_ok) / len(margins_ok), 2) if margins_ok else "",
                round(min(margins_all), 2) if margins_all else "",
                round(max(margins_all), 2) if margins_all else "",
            ])

        Exporter._style_summary_sheet(ws3, len(model_stats))

        # ═══════════════════════════════════════
        #  Sheet 4: TOP OKAZJE
        # ═══════════════════════════════════════
        ws4 = wb.create_sheet("🏆 Top Okazje")

        # Top 20 highest margin cars
        with_margin = [c for c in cars if c.has_catalog_price and not c.is_ev and c.status == "OK"]
        top_deals = sorted(with_margin, key=lambda c: c.margin_pct, reverse=True)[:20]

        ws4.append([
            "# Rank", "Dealer", "Model", "Rok", "Cena Sprzedaży",
            "Cena Katalogowa", "Rabat", "Koszt Dealera",
            "Marża PLN", "Marża %", "URL"
        ])

        for rank, car in enumerate(top_deals, 1):
            ws4.append([
                rank,
                car.dealer_short,
                car.title,
                car.year,
                car.sale_price,
                car.catalog_price,
                car.rebate,
                car.dealer_cost,
                car.margin_pln,
                car.margin_pct,
                car.url,
            ])

        Exporter._style_summary_sheet(ws4, len(top_deals))

        # ═══════════════════════════════════════
        #  Sheet 5: ANOMALIE
        # ═══════════════════════════════════════
        anomalies = [c for c in cars if c.status in ("ANOMALIA", "BRAK_CENY_KAT", "DEMO", "UŻYWANE")]
        if anomalies:
            ws5 = wb.create_sheet("⚠️ Anomalie")
            ws5.append([
                "Dealer", "Model", "Rok", "Status",
                "Cena Sprzedaży", "Cena Katalogowa", "Marża %", "URL"
            ])
            for car in anomalies:
                ws5.append([
                    car.dealer_short,
                    car.title,
                    car.year,
                    car.status,
                    car.sale_price,
                    car.catalog_price if car.catalog_price > 0 else "BRAK",
                    car.margin_pct if car.has_catalog_price else "N/A",
                    car.url,
                ])
            Exporter._style_summary_sheet(ws5, len(anomalies))

        # ═══════════════════════════════════════
        #  Sheet 6: ELEKTRY
        # ═══════════════════════════════════════
        evs = [c for c in cars if c.is_ev]
        if evs:
            ws6 = wb.create_sheet("⚡ Elektryki")
            ws6.append([
                "Dealer", "Model", "Rok", "Moc KM",
                "Cena Sprzedaży", "Cena Katalogowa", "Cena 30 dni",
                "Kolor", "URL"
            ])
            for car in evs:
                ws6.append([
                    car.dealer_short,
                    car.title,
                    car.year,
                    car.power_hp,
                    car.sale_price,
                    car.catalog_price if car.catalog_price > 0 else "",
                    car.price_30d if car.price_30d > 0 else "",
                    car.color,
                    car.url,
                ])
            Exporter._style_summary_sheet(ws6, len(evs))

        wb.save(filepath)
        logging.info(f"  📊 XLSX → {filepath} ({wb.sheetnames})")

    @staticmethod
    def _style_main_sheet(ws, row_count):
        """Stylizacja głównego arkusza."""
        if not HAS_XLSX:
            return

        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Row-level styling based on status
        for row_idx in range(2, row_count + 2):
            status_cell = ws.cell(row=row_idx, column=21)  # U = Status
            status = str(status_cell.value or "")

            # Apply row fill based on status
            fill = None
            if status == "ANOMALIA":
                fill = Exporter.RED_FILL
            elif status == "BRAK_CENY_KAT":
                fill = Exporter.GRAY_FILL
            elif status == "ELEKTRYK_OK":
                fill = Exporter.EV_FILL

            if fill:
                for col_idx in range(1, 27):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

            # Margin cell coloring (column T = 20)
            margin_cell = ws.cell(row=row_idx, column=20)
            if isinstance(margin_cell.value, (int, float)):
                if margin_cell.value >= Config.MARGIN_GREAT:
                    margin_cell.fill = Exporter.GREEN_FILL
                elif margin_cell.value >= Config.MARGIN_BAD:
                    margin_cell.fill = Exporter.YELLOW_FILL
                else:
                    margin_cell.fill = Exporter.RED_FILL

        # Currency formatting (N, O, P, Q, R, S = 14-19)
        for row in ws.iter_rows(min_row=2, min_col=14, max_col=19, max_row=row_count + 1):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")

        # Margin % formatting (T = 20)
        for row in ws.iter_rows(min_row=2, min_col=20, max_col=20, max_row=row_count + 1):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal="center")

        # Price changed highlighting (W = 23)
        for row in ws.iter_rows(min_row=2, min_col=23, max_col=23, max_row=row_count + 1):
            for cell in row:
                if cell.value == "TAK":
                    cell.font = Font(bold=True, color="FF0000")

        # Auto column width
        Exporter._auto_width(ws)

        # Freeze + filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    @staticmethod
    def _style_summary_sheet(ws, row_count):
        """Stylizacja arkuszy podsumowania."""
        if not HAS_XLSX:
            return

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Zebra striping
        stripe_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        for row_idx in range(3, row_count + 2, 2):
            for cell in ws[row_idx]:
                cell.fill = stripe_fill

        Exporter._auto_width(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    @staticmethod
    def _auto_width(ws):
        """Auto-fit column widths."""
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col_cells:
                try:
                    cell_len = len(str(cell.value or ""))
                    max_len = max(max_len, cell_len)
                except:
                    pass
            adjusted = min(max_len + 3, 45)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(adjusted, 8)


# ═══════════════════════════════════════════════════════════════════════════════
#  🚀  SILNIK GŁÓWNY
# ═══════════════════════════════════════════════════════════════════════════════

class GoliathEngine:
    """Orkiestruje cały proces: inventory → memory → parse → calculate → merge → export."""

    def __init__(self):
        self.client = HttpClient()
        self.memory = SmartMemory()
        self.collector = InventoryCollector(self.client)
        self.parser = OfferParser(self.client, self.memory)
        self.results: List[CarData] = []
        self.stats = {
            "dealers_scanned": 0,
            "links_found": 0,
            "links_skipped_cache": 0,
            "offers_parsed": 0,
            "offers_skipped_filter": 0,
            "offers_errors": 0,
            "with_catalog_price": 0,
            "evs": 0,
            "demos": 0,
            "price_changes": 0,
            "removed_listings": 0,
        }

    def run(self):
        start_time = time.time()
        logging.info("🚀 GOLIATH v9.0 ULTRA startuje!\n")

        # ═════════════════════════════════════════════
        #  FAZA 1: Zbieranie linków z inventory
        # ═════════════════════════════════════════════
        logging.info(f"{'═' * 60}")
        logging.info(f"  📡  FAZA 1: Skanowanie {len(DEALERS)} salonów")
        logging.info(f"{'═' * 60}\n")

        all_links: List[Tuple[str, str]] = []

        for i, dealer in enumerate(DEALERS, 1):
            self.stats["dealers_scanned"] += 1
            links = self.collector.collect(dealer["url"])
            for link in links:
                all_links.append((link, dealer["name"]))
            logging.info(f"  🏢 [{i:>2}/{len(DEALERS)}] {dealer['name']:<25} → {len(links)} ofert")

        # Deduplikacja
        seen_urls: Set[str] = set()
        unique_links: List[Tuple[str, str]] = []
        for url, name in all_links:
            if url not in seen_urls:
                seen_urls.add(url)
                unique_links.append((url, name))

        total_found = len(unique_links)
        self.stats["links_found"] = total_found
        logging.info(f"\n  📊 Łącznie: {len(all_links)} linków → {total_found} unikalnych ofert\n")

        # ═════════════════════════════════════════════
        #  FAZA 2: Smart Memory Filter
        # ═════════════════════════════════════════════
        logging.info(f"{'═' * 60}")
        logging.info(f"  🧠  FAZA 2: Smart Memory — filtrowanie znanych aut")
        logging.info(f"{'═' * 60}\n")

        # Detect removed listings
        active_urls = set(url for url, _ in unique_links)
        removed = self.memory.get_removed_urls(active_urls)
        self.stats["removed_listings"] = len(removed)
        if removed:
            logging.info(f"  🗑️  {len(removed)} ogłoszeń zniknęło z Otomoto (sprzedane/wycofane)")

        # Filter: skip cached OK ads
        links_to_scan: List[Tuple[str, str]] = []
        cached_cars: List[CarData] = []

        for url, dealer_name in unique_links:
            if self.memory.should_skip(url):
                self.stats["links_skipped_cache"] += 1
                # Reconstruct from cache for export
                entry = self.memory.cache[url]
                # We'll re-add cached data from previous export
            else:
                links_to_scan.append((url, dealer_name))

        skipped = self.stats["links_skipped_cache"]
        logging.info(f"  ✅ {skipped} aut w cache (stabilna marża — pomijam)")
        logging.info(f"  🔬 {len(links_to_scan)} do skanowania\n")

        if not links_to_scan and skipped == 0:
            logging.warning("  🛑 Brak ofert do analizy!")
            return

        # ═════════════════════════════════════════════
        #  FAZA 3: Deep Scraping
        # ═════════════════════════════════════════════
        if links_to_scan:
            logging.info(f"{'═' * 60}")
            logging.info(f"  🔬  FAZA 3: Analiza {len(links_to_scan)} ogłoszeń ({Config.MAX_WORKERS} wątków)")
            logging.info(f"{'═' * 60}\n")

            with ThreadPoolExecutor(max_workers=Config.MAX_WORKERS) as executor:
                future_map = {}
                for url, dealer_name in links_to_scan:
                    future = executor.submit(self.parser.parse, url, dealer_name)
                    future_map[future] = (url, dealer_name)

                for idx, future in enumerate(as_completed(future_map), 1):
                    url, dealer_name = future_map[future]
                    try:
                        car = future.result()
                        if car is None:
                            self.stats["offers_skipped_filter"] += 1
                            continue

                        self.stats["offers_parsed"] += 1
                        self.results.append(car)

                        if car.has_catalog_price:
                            self.stats["with_catalog_price"] += 1
                        if car.is_ev:
                            self.stats["evs"] += 1
                        if car.is_demo or car.vehicle_type in ("demo", "used"):
                            self.stats["demos"] += 1
                        if car.price_changed:
                            self.stats["price_changes"] += 1

                        self._log_car(idx, len(links_to_scan), car)

                    except Exception as e:
                        self.stats["offers_errors"] += 1
                        logging.error(f"  ❌ [{idx}/{len(links_to_scan)}] {url}: {e}")

        # ═════════════════════════════════════════════
        #  FAZA 4: Merge z cache
        # ═════════════════════════════════════════════
        logging.info(f"\n{'═' * 60}")
        logging.info(f"  🔄  FAZA 4: Merge z cache + eksport")
        logging.info(f"{'═' * 60}\n")

        # Load cached results from previous JSON export (if exists)
        final_results = list(self.results)  # Start with newly scraped

        if os.path.exists(Config.OUTPUT_JSON) and skipped > 0:
            try:
                with open(Config.OUTPUT_JSON, "r", encoding="utf-8") as f:
                    old_data = json.load(f)

                # Merge: add cached entries that we skipped
                # data.json is saved as a flat array, not {"cars": [...]}
                cached_items = old_data if isinstance(old_data, list) else old_data.get("cars", [])
                new_urls = set(car.url for car in final_results)
                for item in cached_items:
                    url = item.get("url", "")
                    if url and url not in new_urls and self.memory.should_skip(url):
                        # Reconstruct CarData from JSON
                        cached_car = CarData(
                            otomoto_id=item.get("otomoto_id", ""),
                            dealer=item.get("dealer", ""),
                            dealer_short=item.get("dealer_short", ""),
                            title=item.get("title", ""),
                            model=item.get("model", ""),
                            model_raw=item.get("model_raw", ""),
                            year=item.get("year", 0),
                            fuel=item.get("fuel", ""),
                            power_hp=item.get("power_hp", 0),
                            engine_capacity=item.get("engine_capacity", ""),
                            mileage_km=item.get("mileage_km", 0),
                            gearbox=item.get("gearbox", ""),
                            drive=item.get("drive", ""),
                            color=item.get("color", ""),
                            color_type=item.get("color_type", ""),
                            body_type=item.get("body_type", ""),
                            doors=item.get("doors", 0),
                            seats=item.get("seats", 0),
                            is_new=item.get("is_new", True),
                            has_vin=item.get("has_vin", False),
                            sale_price=item.get("sale_price", 0),
                            catalog_price=item.get("catalog_price", 0),
                            price_30d=item.get("price_30d", 0),
                            dealer_cost=item.get("dealer_cost", 0),
                            description=item.get("description", ""),
                            rebate=item.get("rebate", 0),
                            rebate_applied=item.get("rebate_applied", False),
                            margin_pln=item.get("margin_pln", 0),
                            margin_pct=item.get("margin_pct", 0.0),
                            vehicle_type=item.get("vehicle_type", "new"),
                            is_demo=item.get("is_demo", False),
                            has_catalog_price=item.get("has_catalog_price", False),
                            is_ev=item.get("is_ev", False),
                            status=item.get("status", ""),
                            url=url,
                            scraped_at=item.get("scraped_at", ""),
                            first_seen=item.get("first_seen", ""),
                            price_changed=False,
                        )
                        # Only add if still active on Otomoto
                        if url in active_urls:
                            final_results.append(cached_car)
                            new_urls.add(url)

                logging.info(f"  📦 Merged: {len(final_results)} aut ({len(self.results)} nowych + {len(final_results) - len(self.results)} z cache)")

            except Exception as e:
                logging.warning(f"  ⚠️ Nie udało się merge'ować: {e}")

        # Sort
        status_order = {"OK": 0, "DEMO": 1, "UŻYWANE": 2, "ELEKTRYK_OK": 3, "ANOMALIA": 4, "BRAK_CENY_KAT": 5}
        final_results.sort(
            key=lambda c: (status_order.get(c.status, 9), -c.margin_pct)
        )

        # Export
        if final_results:
            Exporter.to_json(final_results, Config.OUTPUT_JSON)
            Exporter.to_xlsx(final_results, Config.OUTPUT_XLSX)
            self.memory.save(final_results)
        else:
            logging.warning("  🛑 Brak danych do eksportu.")

        # ═════════════════════════════════════════════
        #  PODSUMOWANIE
        # ═════════════════════════════════════════════
        elapsed = round((time.time() - start_time) / 60, 2)
        self._print_summary(elapsed, final_results)

    def _log_car(self, idx: int, total: int, car: CarData):
        prefix = f"  [{idx:>3}/{total}]"

        if car.is_ev:
            logging.info(
                f"{prefix} ⚡ {car.dealer_short:<20} │ "
                f"{car.title[:28]:<28} │ {car.year} │ "
                f"Sprzedaż: {car.sale_price:>9,} │ ELEKTRYK"
                + (f" │ Katalog: {car.catalog_price:>9,}" if car.has_catalog_price else "")
            )
        elif car.has_catalog_price:
            if car.margin_pct >= Config.MARGIN_GREAT:
                icon = "🟢"
            elif car.margin_pct >= Config.MARGIN_BAD:
                icon = "✅"
            else:
                icon = "🔴"

            change = " 📈" if car.price_changed else ""
            logging.info(
                f"{prefix} {icon} {car.dealer_short:<20} │ "
                f"{car.title[:28]:<28} │ {car.year} │ "
                f"Sprzedaż: {car.sale_price:>9,} │ "
                f"Katalog: {car.catalog_price:>9,} │ "
                f"Marża: {car.margin_pct:>6.1f}% ({car.margin_pln:>+8,} PLN){change}"
            )
        else:
            logging.info(
                f"{prefix} ❓ {car.dealer_short:<20} │ "
                f"{car.title[:28]:<28} │ {car.year} │ "
                f"Sprzedaż: {car.sale_price:>9,} │ BRAK CENY KATALOGOWEJ"
            )

    def _print_summary(self, elapsed_min: float, all_cars: List[CarData]):
        s = self.stats

        logging.info(f"""
{'═' * 60}
  🏆  GOLIATH v9.0 ULTRA — RAPORT KOŃCOWY
{'═' * 60}

  🏢 Salonów przeskanowanych:    {s['dealers_scanned']:>5}
  🔗 Ofert na Otomoto:           {s['links_found']:>5}
  🧠 Pominięte (cache OK):       {s['links_skipped_cache']:>5}
  🔬 Przeanalizowanych:          {s['offers_parsed']:>5}
  📊 Z ceną katalogową:          {s['with_catalog_price']:>5}
  ⚡ Elektryków:                 {s['evs']:>5}
  🏷️  Demo/Używanych:             {s['demos']:>5}
  ⏭️  Odfiltrowanych:             {s['offers_skipped_filter']:>5}
  ❌ Błędów:                     {s['offers_errors']:>5}
  📈 Zmian cen:                  {s['price_changes']:>5}
  🗑️  Znikniętych ogłoszeń:      {s['removed_listings']:>5}
  🌐 Requestów HTTP:             {self.client.total_requests:>5}
  ⏱️  Czas operacji:              {elapsed_min:.1f} min

  📦 TOTAL w bazie:              {len(all_cars):>5} aut

{'═' * 60}""")

        # Margin analysis
        # Status counts
        status_counts = {}
        for car in all_cars:
            status_counts[car.status] = status_counts.get(car.status, 0) + 1

        logging.info(f"""
  📋 STATUSY:
     ✅ OK:            {status_counts.get('OK', 0):>5}
     🏷️  DEMO:          {status_counts.get('DEMO', 0):>5}
     🔄 UŻYWANE:       {status_counts.get('UŻYWANE', 0):>5}
     ⚡ ELEKTRYK_OK:   {status_counts.get('ELEKTRYK_OK', 0):>5}
     ⚠️  ANOMALIA:      {status_counts.get('ANOMALIA', 0):>5}
     ❓ BRAK_CENY_KAT: {status_counts.get('BRAK_CENY_KAT', 0):>5}
""")

        with_margin = [c for c in all_cars if c.status == "OK"]
        if with_margin:
            avg = sum(c.margin_pct for c in with_margin) / len(with_margin)
            best = max(with_margin, key=lambda c: c.margin_pct)
            worst = min(with_margin, key=lambda c: c.margin_pct)

            logging.info(f"""
  📈 ANALIZA MARŻY ({len(with_margin)} ofert ze statusem OK, zakres -0.5% do 6%):

     Średnia marża:  {avg:>6.1f}%
     🟢 Najwyższa:  {best.margin_pct:>6.1f}%  — {best.title[:35]} ({best.dealer_short})
     🔴 Najniższa:  {worst.margin_pct:>6.1f}%  — {worst.title[:35]} ({worst.dealer_short})
""")

        # Model breakdown — ONLY OK status
        model_stats: Dict[str, List[float]] = {}
        for car in with_margin:
            model = car.model or car.title
            model_stats.setdefault(model, []).append(car.margin_pct)

        if model_stats:
            logging.info("  📋 MARŻA WG MODELU:")
            logging.info(f"  {'Model':<25} {'Ilość':>6} {'Śr. marża':>10} {'Min':>8} {'Max':>8}")
            logging.info(f"  {'─' * 60}")

            for model, margins in sorted(model_stats.items()):
                avg_m = sum(margins) / len(margins)
                logging.info(
                    f"  {model:<25} {len(margins):>6} {avg_m:>9.1f}% {min(margins):>7.1f}% {max(margins):>7.1f}%"
                )

        # Dealer ranking — ONLY OK status
        dealer_avgs: Dict[str, Tuple[float, int]] = {}
        for car in with_margin:
            d = car.dealer_short
            if d not in dealer_avgs:
                dealer_avgs[d] = (0.0, 0)
            total_m, count = dealer_avgs[d]
            dealer_avgs[d] = (total_m + car.margin_pct, count + 1)

        if dealer_avgs:
            logging.info(f"\n  🏅 RANKING DEALERÓW (wg średniej marży):")
            logging.info(f"  {'Dealer':<25} {'Aut':>6} {'Śr. marża':>10}")
            logging.info(f"  {'─' * 45}")

            sorted_dealers = sorted(
                dealer_avgs.items(),
                key=lambda x: x[1][0] / x[1][1],
                reverse=True
            )
            for dealer_name, (total_m, count) in sorted_dealers:
                avg_m = total_m / count
                medal = "🥇" if avg_m == max(t/c for t,c in dealer_avgs.values()) else "  "
                logging.info(
                    f"  {medal}{dealer_name:<23} {count:>6} {avg_m:>9.1f}%"
                )

        # Smart Rebate summary
        rebate_applied_count = sum(1 for c in all_cars if c.rebate_applied)
        rebate_not_applied = sum(1 for c in all_cars if c.has_catalog_price and not c.is_ev and not c.rebate_applied)
        logging.info(f"""
  💰 SMART REBATE:
     Rabat zastosowany:      {rebate_applied_count:>5} aut
     Rabat niepotrzebny:     {rebate_not_applied:>5} aut (marża OK bez rabatu)
""")

        logging.info(f"""
  ✅ Pliki wyjściowe:
     {Config.OUTPUT_JSON}
     {Config.OUTPUT_XLSX}
     {Config.CACHE_FILE}

{'═' * 60}
""")


# ═══════════════════════════════════════════════════════════════════════════════
#  🎬  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def setup_logging():
    log_format = "%(asctime)s │ %(message)s"
    date_format = "%H:%M:%S"

    file_handler = logging.FileHandler(Config.LOG_FILE, encoding="utf-8", mode="w")
    file_handler.setFormatter(logging.Formatter(log_format, datefmt=date_format))
    file_handler.setLevel(logging.DEBUG)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(logging.Formatter(log_format, datefmt=date_format))
    console_handler.setLevel(logging.INFO)

    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)


def main():
    # Fix Windows encoding (cp1250 -> utf-8) to prevent UnicodeEncodeError
    import sys, io
    try:
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
            sys.stderr.reconfigure(encoding='utf-8', errors='replace')
        else:
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
            sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass

    setup_logging()

    print("""
    ╔══════════════════════════════════════════════════════════════╗
    ║                                                              ║
    ║   ██████╗  ██████╗ ██╗     ██╗ █████╗ ████████╗██╗  ██╗    ║
    ║  ██╔════╝ ██╔═══██╗██║     ██║██╔══██╗╚══██╔══╝██║  ██║    ║
    ║  ██║  ███╗██║   ██║██║     ██║███████║   ██║   ███████║    ║
    ║  ██║   ██║██║   ██║██║     ██║██╔══██║   ██║   ██╔══██║    ║
    ║  ╚██████╔╝╚██████╔╝███████╗██║██║  ██║   ██║   ██║  ██║    ║
    ║   ╚═════╝  ╚═════╝ ╚══════╝╚═╝╚═╝  ╚═╝   ╚═╝   ╚═╝  ╚═╝    ║
    ║                                                              ║
    ║  v9.0 ULTRA │ 29 DEALERÓW │ SMART REBATE │ DEMO DETECTION    ║
    ║                                                              ║
    ╚══════════════════════════════════════════════════════════════╝
    """)

    engine = GoliathEngine()
    engine.run()


if __name__ == "__main__":
    main()
